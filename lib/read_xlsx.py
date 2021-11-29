import openpyxl
import pandas as pd
from pandas.core.series import Series
from pandas.core.frame import DataFrame
from typing import List

def preprocess_xlsx(xlsx_fnames:List[str]) -> DataFrame:
    dflist = []
    for xlsx_fname in xlsx_fnames:
        wb = openpyxl.load_workbook(xlsx_fname)    
        for sheet_name in wb.sheetnames:
            sheet= wb[sheet_name]
            for m in sheet.merged_cell_ranges:
                # unmerge cells
                sheet.unmerge_cells(
                    start_column=m.min_col,
                    start_row=m.min_row,
                    end_column=m.max_col,
                    end_row=m.max_row,
                )
                value = sheet.cell(row=m.min_row, column=m.min_col).value
                for col in range(m.min_col, m.max_col+1):
                    for row in range(m.min_row, m.max_row+1):
                        sheet.cell(row=row, column=col).value = value
            
            # set column
            tmpdf = pd.DataFrame(sheet.values).dropna(how='all')
            tmpdf.columns = ['day','dow','class'] + tmpdf.iloc[1][3:].tolist()

            # add column [year, month]
            srs_ym = tmpdf['day'].str.split('   授業時間割').str[0]
            tmpdf['year'] = srs_ym.str.split('（').str[0].ffill()
            tmpdf['month'] = srs_ym.str.split('年').str[1].str[:-1].ffill()
            
            # drop title-row
            tmpdf = tmpdf[srs_ym.isna()]
            # drop header-row
            tmpdf = tmpdf[~tmpdf[['day','class']].isna().any(1)]
            # drop duplicates-columns
            tmpdf2 = pd.DataFrame(columns=['year', 'month', 'day', '1限\n9:20 - 11:00', '2限\n11:10 - 12:50', 
                    '3限\n13:50 - 15:30','4限\n15:40 - 17:20', '5限\n17:30 - 19:10', '6限\n18:30 - 20:10', '7限\n20:15 - 21:55'])
            for col in tmpdf2.columns:
                if type(tmpdf[col])==Series:
                    tmpdf2[col] = tmpdf[col]
                else:
                    tmpdf2[col] = tmpdf[col].apply(set,axis=1).apply(list).str[0]
            dflist.append(tmpdf2)
    df = pd.concat(dflist)
    df['year'] = df['year'].astype(int)
    df['month'] = df['month'].astype(int)
    df = df.drop_duplicates().set_index(['year', 'month', 'day'])
    df.columns = [1, 2, 3, 4, 5, 6, 7]
    return df


def get_unique_lecture(df: DataFrame) -> List[str]:
    leclist = []
    for col in df.columns:
        lecs = df[col].dropna().unique()
        leclist.extend(lecs)
    leclist = sorted(list(set(leclist)))

    leclist = [lec for lec in leclist if '「' in lec]
    return leclist